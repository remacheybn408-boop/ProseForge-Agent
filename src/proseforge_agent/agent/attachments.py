"""Deterministic attachment ingestion for chat and project context."""

from __future__ import annotations

import csv
import hashlib
import json
import struct
import zipfile
from dataclasses import asdict, dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Callable
from xml.etree import ElementTree

from ..chat.transcript import append_jsonl
from ..errors import ConfigurationError
from ..novel.artifacts import ArtifactGraphStore, ArtifactRecord


TEXT_EXTENSIONS = {".txt": "text", ".md": "markdown", ".markdown": "markdown"}
TABLE_EXTENSIONS = {".csv": "csv", ".xlsx": "excel", ".xlsm": "excel"}
DOCUMENT_EXTENSIONS = {".pdf": "pdf", ".docx": "docx"}
IMAGE_EXTENSIONS = {".png": "image", ".jpg": "image", ".jpeg": "image", ".webp": "image"}
SUPPORTED_EXTENSIONS = set(TEXT_EXTENSIONS) | set(TABLE_EXTENSIONS) | set(DOCUMENT_EXTENSIONS) | set(IMAGE_EXTENSIONS)


@dataclass(frozen=True)
class AttachmentIngestResult:
    """One ingested attachment and its persisted project artifacts."""

    id: str
    status: str
    kind: str
    source_path: Path
    text: str = ""
    metadata: dict[str, Any] = field(default_factory=dict)
    artifact_path: Path = Path()
    searchable_path: Path | None = None
    memory_candidate_path: Path | None = None
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        payload["source_path"] = str(self.source_path)
        payload["artifact_path"] = str(self.artifact_path)
        payload["searchable_path"] = str(self.searchable_path) if self.searchable_path else None
        payload["memory_candidate_path"] = str(self.memory_candidate_path) if self.memory_candidate_path else None
        return payload


class AttachmentIngestor:
    """Extract searchable text and metadata from supported attachments."""

    def __init__(
        self,
        root: str | Path,
        *,
        vision_describer: Callable[[Path], str] | None = None,
    ) -> None:
        self.root = Path(root)
        self.vision_describer = vision_describer

    def ingest_file(self, source: str | Path, *, slug: str) -> AttachmentIngestResult:
        path = Path(source)
        self._require_source(path)
        if path.suffix.lower() in IMAGE_EXTENSIONS:
            return self.ingest_image(path, slug=slug)
        text, kind, metadata, warnings = self._extract(path)
        return self._persist(path, slug=slug, kind=kind, text=text, metadata=metadata, warnings=warnings)

    def ingest_image(self, source: str | Path, *, slug: str) -> AttachmentIngestResult:
        path = Path(source)
        self._require_source(path)
        metadata = _image_metadata(path)
        text = self.vision_describer(path) if self.vision_describer else ""
        warnings = [] if text else ["image description provider not configured"]
        return self._persist(path, slug=slug, kind="image", text=text, metadata=metadata, warnings=warnings)

    def ingest_folder(self, source: str | Path, *, slug: str) -> list[AttachmentIngestResult]:
        folder = Path(source)
        if not folder.is_dir():
            raise ConfigurationError(f"attachment folder does not exist: {folder}")
        results: list[AttachmentIngestResult] = []
        for item in sorted(folder.iterdir(), key=lambda path: path.name):
            if item.is_file() and item.suffix.lower() in SUPPORTED_EXTENSIONS:
                results.append(self.ingest_file(item, slug=slug))
        return results

    def _extract(self, path: Path) -> tuple[str, str, dict[str, Any], list[str]]:
        suffix = path.suffix.lower()
        metadata = {"extension": suffix, "size_bytes": path.stat().st_size}
        if suffix in TEXT_EXTENSIONS:
            return path.read_text(encoding="utf-8"), TEXT_EXTENSIONS[suffix], metadata, []
        if suffix == ".csv":
            return _extract_csv(path), "csv", metadata, []
        if suffix in {".xlsx", ".xlsm"}:
            text, warnings = _extract_excel(path)
            return text, "excel", metadata, warnings
        if suffix == ".docx":
            return _extract_docx(path), "docx", metadata, []
        if suffix == ".pdf":
            text, warnings = _extract_pdf(path)
            return text, "pdf", metadata, warnings
        raise ConfigurationError(f"unsupported attachment format: {suffix}")

    def _persist(
        self,
        source_path: Path,
        *,
        slug: str,
        kind: str,
        text: str,
        metadata: dict[str, Any],
        warnings: list[str],
    ) -> AttachmentIngestResult:
        project_root = self.root / "projects" / slug
        attachment_id = f"attachment_{_sha256(source_path.read_bytes())[:12]}"
        artifact_path = project_root / "attachments" / f"{attachment_id}.json"
        searchable_path = project_root / "searchable" / f"{attachment_id}.txt"
        candidate_path = self.root / "memory_candidates" / "projects" / f"{slug}.jsonl"
        searchable_text = text if text.strip() else _metadata_text(metadata)

        artifact_path.parent.mkdir(parents=True, exist_ok=True)
        searchable_path.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "id": attachment_id,
            "kind": kind,
            "source": str(source_path),
            "text": text,
            "metadata": metadata,
            "warnings": warnings,
            "created_at": datetime.now(UTC).isoformat(),
        }
        artifact_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        searchable_path.write_text(searchable_text, encoding="utf-8")
        append_jsonl(
            candidate_path,
            {
                "id": f"memcand_{attachment_id}",
                "kind": "attachment",
                "project_slug": slug,
                "scope": "project",
                "source": str(source_path),
                "text": searchable_text[:2000],
                "status": "candidate",
                "created_at": datetime.now(UTC).isoformat(),
            },
        )
        ArtifactGraphStore(self.root, slug=slug).add(
            ArtifactRecord(
                id=attachment_id,
                type=f"attachment:{kind}",
                generated=[str(searchable_path)],
                checksum=_sha256(artifact_path.read_bytes()),
                provider="local",
                prompt_version="attachment-ingestion-v1",
            )
        )
        return AttachmentIngestResult(
            id=attachment_id,
            status="degraded" if warnings and not text.strip() else "ok",
            kind=kind,
            source_path=source_path,
            text=text,
            metadata=metadata,
            artifact_path=artifact_path,
            searchable_path=searchable_path,
            memory_candidate_path=candidate_path,
            warnings=warnings,
        )

    @staticmethod
    def _require_source(path: Path) -> None:
        if not path.is_file():
            raise ConfigurationError(f"attachment file does not exist: {path}")


def _extract_csv(path: Path) -> str:
    rows: list[str] = []
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            rows.append("\t".join(row))
    return "\n".join(rows) + ("\n" if rows else "")


def _extract_docx(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        try:
            raw = archive.read("word/document.xml")
        except KeyError:
            return ""
    root = ElementTree.fromstring(raw)
    namespace = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    texts = [node.text or "" for node in root.iter(f"{namespace}t")]
    return "\n".join(text for text in texts if text)


def _extract_pdf(path: Path) -> tuple[str, list[str]]:
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception:
        try:
            from PyPDF2 import PdfReader  # type: ignore
        except Exception:
            return "", ["pdf text extraction requires pypdf or PyPDF2"]
    try:
        reader = PdfReader(str(path))
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
    except Exception as exc:
        return "", [f"pdf text extraction failed: {exc}"]
    return text, []


def _extract_excel(path: Path) -> tuple[str, list[str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return "", ["excel table extraction requires openpyxl"]
    workbook = load_workbook(path, read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        lines.append(f"# {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value) for value in row]
            if any(values):
                lines.append("\t".join(values))
    return "\n".join(lines) + ("\n" if lines else ""), []


def _image_metadata(path: Path) -> dict[str, Any]:
    data = path.read_bytes()
    metadata: dict[str, Any] = {
        "extension": path.suffix.lower(),
        "size_bytes": len(data),
    }
    if data.startswith(b"\x89PNG\r\n\x1a\n") and len(data) >= 24:
        width, height = struct.unpack(">II", data[16:24])
        metadata.update({"format": "png", "width": width, "height": height})
        return metadata
    if data.startswith(b"\xff\xd8"):
        size = _jpeg_size(data)
        metadata.update({"format": "jpeg"})
        if size:
            metadata.update({"width": size[0], "height": size[1]})
        return metadata
    metadata["format"] = path.suffix.lower().lstrip(".") or "unknown"
    return metadata


def _jpeg_size(data: bytes) -> tuple[int, int] | None:
    index = 2
    while index + 9 < len(data):
        if data[index] != 0xFF:
            index += 1
            continue
        marker = data[index + 1]
        block_length = int.from_bytes(data[index + 2 : index + 4], "big")
        if marker in {0xC0, 0xC2}:
            height = int.from_bytes(data[index + 5 : index + 7], "big")
            width = int.from_bytes(data[index + 7 : index + 9], "big")
            return width, height
        index += 2 + block_length
    return None


def _metadata_text(metadata: dict[str, Any]) -> str:
    return "\n".join(f"{key}: {value}" for key, value in sorted(metadata.items())) + "\n"


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


__all__ = [
    "AttachmentIngestResult",
    "AttachmentIngestor",
    "SUPPORTED_EXTENSIONS",
]
